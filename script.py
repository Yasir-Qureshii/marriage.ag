import os.path
import openpyxl
import time
import re
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select, WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import StaleElementReferenceException
from bs4 import BeautifulSoup

pattern = 'Religious marriage celebrant'
ph = 'p(H):'
pw = 'p(W):'
m = 'm:'

# Create new file if already exists    
filename = 'MinistersOfReligion.xlsx'
file_exists = os.path.exists(filename)

if file_exists:
    filename = filename.replace('.xlsx', ' 2.xlsx')

# Initialize the driver
options = Options()
options.headless = True
driver = webdriver.Chrome(options=options, service=Service(ChromeDriverManager().install()))
url = 'https://marriage.ag.gov.au/statecelebrants/other'
driver.get(url)
# driver.maximize_window()
time.sleep(3)
# Get states from the page
states = []
states_selectors = driver.find_elements(By.XPATH, '//*[@id="ctl00_MainContent_rmnStates"]/ul/li[1]/following-sibling::li')
for state in states_selectors:
    states.append(state.text.strip())
    

# Create workbook and add sheet for each state
header = ['SURNAME', 'First Name', 'Salutation', 'Date', 'Status', 'Religious?', 'Address 1', 'City', 'State', 'Postcode', 'p(H)', 'p(W)', 'm', 'email', 'Denomination']
excel = openpyxl.Workbook()
sheet1 = excel.active
sheet1.title = states[0]
sheet1.append(header)
for cell in sheet1["1:1"]:
    cell.font = Font(bold=True)
        
for state in states[1:]:
    ws = excel.create_sheet()
    ws.title = state
    ws.append(header)
    for cell in ws["1:1"]:
        cell.font = Font(bold=True)

excel.save(filename)


# Function Data Scraper
def scrape_table(driver, sheet, page, state, elem):
    global excel, pattern, ph, pw, m, filename
    soup = BeautifulSoup(elem.get_attribute('outerHTML'), "lxml")
    
    trs = soup.find_all('tr', id=True)
    for tr in trs:
        tds = tr.find_all('td')
        td1 = tds[2]
        td2 = tds[3]
        spans_td2 = td2.find_all('span')
        surname = td1.find('b').text
        name_ = td1.find('span').contents[-1].split(',')
        first_name  = name_[0]
        if len(name_) > 2:
            salutation = name_[1] + ' ' + name_[2]
        else:
            salutation = name_[1]
        
        salutation = salutation.strip()
        business = td2.find('b').text
        address = td2.find('span').contents
        
        city = None
        postcode = None
        if len(address) > 2:
            address2 = address[-1].split(',')
            city = address2[0].replace(state, '').strip()
            postcode = address2[-1].strip()
            
            address1 = address[0] + ' ' + address[-1]
            address1 = address1.strip()
        else:
            address1 = td2.find('span').text.strip()
            if state in address1:
                city = address1.split(',')[0].replace(state, '').strip()
                postcode = address1.split(',')[-1].strip()
                
        
        # Religion	SURNAME	First Name	Salutation	Address 1	City	State	Postcode	Business Name
        row = [business, surname, first_name, salutation, address1, city, state, postcode, business]
        print(page, ' : ', row)
        sheet.append(row)
    excel.save(filename)


def next_exists(driver):
    try:
        driver.find_element(By.CLASS_NAME, 'rgPageNext').click()
        time.sleep(3)
        return 'True'
    except ElementClickInterceptedException as error:
        print('error : ElementClickInterceptedException')
        return 'False'
    except StaleElementReferenceException as error2:
        print('error : StaleElementReferenceException')
        return 'False'
    
    
# Click each state tab and scrape the data
for state_selector in states_selectors:
    # Switch to Current State in excel
    state = state_selector.text.strip()
    excel.active = excel[state]
    sheet = excel.active
    
    # Switch to Current State in browser
    state_selector.find_element(By.TAG_NAME, 'a').click()
    time.sleep(5)
    
    # Get total pages for current state
    total_pages = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.XPATH, '//*[@id="ctl00_MainContent_gridCelebrants_ctl00"]/thead/tr[1]/td/table/tbody/tr/td/div[5]/strong[2]'))
    ).text.strip()
    
    print('total_pages: ', total_pages)
    
    current_page = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CLASS_NAME, 'rgCurrentPage'))
    ).text.strip()
    
    while int(current_page) <= int(total_pages):
        elem_selector = driver.find_element(By.XPATH, '//*[@id="ctl00_MainContent_gridCelebrants_ctl00"]/tbody')
        scrape_table(driver, sheet, current_page, state, elem_selector)
        page_loaded = next_exists(driver)
        while page_loaded == 'False':
            page_loaded = next_exists(driver)
                
        current_page = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CLASS_NAME, 'rgCurrentPage'))
                ).text.strip()
            
        # Break if its last page
        if int(current_page) == int(total_pages):
            break
        
    # Scrape the last page
    elem_selector = driver.find_element(By.XPATH, '//*[@id="ctl00_MainContent_gridCelebrants_ctl00"]/tbody')
    scrape_table(driver, sheet, current_page, state, elem_selector)


driver.quit()
